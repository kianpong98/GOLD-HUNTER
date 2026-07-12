#!/usr/bin/env python3
"""Update central-bank gold reserve holdings without touching any news data.

Sources are World Gold Council official XLSX files compiled from IMF IFS.
Existing successful data is never erased. A first sync with no usable records fails
visibly instead of producing a misleading green workflow.
"""
from __future__ import annotations

import io, json, re, sys, time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "assets" / "data" / "central-bank-gold-reserves.json"
PAGE = "https://www.gold.org/goldhub/data/gold-reserves-by-country"
# Current official files exposed by the WGC page. Discovery still runs first so later
# monthly file changes can be picked up automatically.
DIRECT_FALLBACKS = [
    "https://www.gold.org/download/file/7739/World_official_gold_holdings_%20as%20of%20Jul2026_IFS.xlsx",
    "https://china.gold.org/download/file/18566/%E5%85%A8%E7%90%83%E5%AE%98%E6%96%B9%E9%BB%84%E9%87%91%E5%82%A8%E5%A4%87.xlsx",
]
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": PAGE,
}
COUNTRY_ALIASES = {
    "united states": "United States", "usa": "United States", "u.s.": "United States",
    "china": "China", "people's republic of china": "China",
    "india": "India", "turkey": "Türkiye", "türkiye": "Türkiye",
    "russia": "Russia", "russian federation": "Russia", "poland": "Poland",
    "singapore": "Singapore", "kazakhstan": "Kazakhstan", "malaysia": "Malaysia",
    "germany": "Germany", "italy": "Italy", "france": "France", "switzerland": "Switzerland",
}
PRIORITY = ["United States", "China", "India", "Türkiye", "Russia", "Poland", "Singapore", "Kazakhstan", "Malaysia"]


def load_existing() -> dict[str, Any]:
    try: return json.loads(OUT.read_text(encoding="utf-8"))
    except Exception: return {"records": [], "history": [], "updatedAt": None}


def num(v: Any) -> float | None:
    if isinstance(v, (int, float)): return float(v)
    s = str(v or "").strip().replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else None


def normalize_country(v: Any) -> str | None:
    raw = re.sub(r"\s+", " ", str(v or "").strip())
    return COUNTRY_ALIASES.get(raw.casefold(), raw) if raw else None


def get(session: requests.Session, url: str, *, binary=False) -> requests.Response:
    last = None
    for attempt in range(4):
        try:
            r = session.get(url, headers=HEADERS, timeout=60, allow_redirects=True)
            if r.status_code == 200 and (not binary or len(r.content) > 5000): return r
            last = RuntimeError(f"HTTP {r.status_code}, {len(r.content)} bytes")
        except Exception as exc: last = exc
        time.sleep(2 ** attempt)
    raise RuntimeError(f"request failed for {url}: {last}")


def discover_downloads(session: requests.Session) -> list[str]:
    html = get(session, PAGE).text.replace("\\/", "/")
    # Capture normal hrefs and links embedded in JSON/Drupal state.
    raw = re.findall(r'''(?:href|url|downloadUrl)\s*["']?\s*[:=]\s*["']([^"']+\.(?:xlsx|xls|csv)(?:\?[^"']*)?)["']''', html, re.I)
    raw += re.findall(r'''https?://[^"'<>\s]+\.(?:xlsx|xls|csv)(?:\?[^"'<>\s]*)?''', html, re.I)
    urls=[]
    for link in raw + DIRECT_FALLBACKS:
        u=urljoin(PAGE, link)
        if u not in urls: urls.append(u)
    return sorted(urls, key=lambda u: sum(k in u.lower() for k in ("world_official", "holding", "reserve", "official")), reverse=True)


def parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    best=[]
    for ws in wb.worksheets:
        rows=[list(row) for row in ws.iter_rows(values_only=True)]
        parsed=parse_rows(rows)
        if len(parsed)>len(best): best=parsed
    return best


def parse_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    best=[]
    for header_idx in range(min(80,len(rows))):
        labels=[str(x or "").strip().casefold() for x in rows[header_idx]]
        country_cols=[j for j,s in enumerate(labels) if any(k in s for k in ("country", "economy", "institution"))]
        holding_cols=[j for j,s in enumerate(labels) if any(k in s for k in ("tonnes", "tonne", "metric ton")) or s in {"gold", "holdings"}]
        if not country_cols or not holding_cols: continue
        # Prefer the first tonnes column after the country column.
        ci=country_cols[0]
        hi=next((j for j in holding_cols if j>ci), holding_cols[0])
        share=next((j for j,s in enumerate(labels) if "%" in s and "reserve" in s),None)
        period=next((j for j,s in enumerate(labels) if any(k in s for k in ("holdings as of","date","month","period"))),None)
        out=[]
        for row in rows[header_idx+1:]:
            if len(row)<=max(ci,hi): continue
            country=normalize_country(row[ci]); holdings=num(row[hi])
            if not country or holdings is None or not (0 < holdings < 20000): continue
            item={"country":country,"holdingsTonnes":round(holdings,2)}
            if share is not None and share<len(row):
                v=num(row[share])
                if v is not None and 0<=v<=100: item["shareOfReservesPct"]=round(v,2)
            if period is not None and period<len(row) and row[period] is not None: item["period"]=str(row[period]).strip()
            out.append(item)
        dedup={x["country"]:x for x in out}
        if len(dedup)>len(best): best=list(dedup.values())
    return best


def main() -> int:
    existing=load_existing(); errors=[]
    session=requests.Session()
    try:
        # Establish WGC cookies first; failure is tolerated because direct files may work.
        try: get(session,PAGE)
        except Exception as e: errors.append(str(e))
        candidates=discover_downloads(session)
        chosen_url=None; records=[]
        for url in candidates:
            try:
                r=get(session,url,binary=True)
                if r.content[:2] != b"PK": raise RuntimeError("response is not an XLSX file")
                parsed=parse_xlsx(r.content)
                selected=[next((x for x in parsed if x["country"]==name),None) for name in PRIORITY]
                selected=[x for x in selected if x]
                if len(selected)>=5:
                    chosen_url=url; records=selected; break
                errors.append(f"{url}: only {len(selected)} priority countries parsed")
            except Exception as e: errors.append(f"{url}: {e}")
        if not records: raise RuntimeError(" | ".join(errors[-6:]) or "No official WGC file could be parsed")
        updated=datetime.now(timezone.utc).isoformat().replace("+00:00","Z")
        # Preserve manually/previously known monthly changes when holdings refresh has none.
        old_by={x.get("country"):x for x in existing.get("records",[]) if isinstance(x,dict)}
        for x in records:
            old=old_by.get(x["country"],{})
            if "monthlyChangeTonnes" not in x and isinstance(old.get("monthlyChangeTonnes"),(int,float)):
                x["monthlyChangeTonnes"]=old["monthlyChangeTonnes"]
        changes=[x.get("monthlyChangeTonnes") for x in records if isinstance(x.get("monthlyChangeTonnes"),(int,float))]
        payload={
            "source":"World Gold Council / IMF IFS","sourceUrl":PAGE,"downloadUrl":chosen_url,"updatedAt":updated,
            "records":records,
            "summary":{"countriesTracked":len(records),"netMonthlyChangeTonnes":round(sum(changes),2) if changes else None,
                       "signal":"Net Buying" if changes and sum(changes)>0 else "Net Selling" if changes and sum(changes)<0 else "Monthly change unavailable"},
            "status":"ready"
        }
        OUT.parent.mkdir(parents=True,exist_ok=True)
        OUT.write_text(json.dumps(payload,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
        print(f"SYNC_OK records={len(records)} source={chosen_url}")
        return 0
    except Exception as exc:
        print(f"SYNC_FAILED: {exc}",file=sys.stderr)
        # Never erase a successful snapshot. But first sync must visibly fail.
        if existing.get("records"):
            print(f"Keeping existing snapshot with {len(existing['records'])} records.")
            return 0
        return 2

if __name__=="__main__": raise SystemExit(main())
