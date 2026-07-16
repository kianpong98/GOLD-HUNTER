#!/usr/bin/env python3
"""Synchronise SPDR GLD tonnes from the official historical XLSX archive.

The official archive is the source of truth. The script is deliberately
non-destructive: a failed or stale fetch never replaces an existing good file,
and an unchanged archive does not create a new Git commit/deployment.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import time
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

API = "https://api.spdrgoldshares.com/api/v1/historical-archive?exchange=NYSE&lang=en&product=gld"
PAGE = "https://www.spdrgoldshares.com/usa/gld/"
OUT = Path("assets/data/spdr-gld-holdings.json")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131 Safari/537.36",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/json,text/html;q=0.8,*/*;q=0.5",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": PAGE,
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}


def session() -> requests.Session:
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        backoff_factor=2,
        status_forcelist=(429, 500, 502, 503, 504, 520, 522, 524),
        allowed_methods=frozenset({"GET"}),
        respect_retry_after_header=True,
    )
    s = requests.Session()
    s.headers.update(HEADERS)
    s.mount("https://", HTTPAdapter(max_retries=retry))
    return s


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    for fmt in (
        "%Y-%m-%d",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%b %d, %Y",
        "%B %d, %Y",
    ):
        try:
            return datetime.strptime(text[:30], fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return None


def looks_like_xlsx(content: bytes) -> bool:
    return content.startswith(b"PK\x03\x04")


def find_download(value: Any) -> str | None:
    if isinstance(value, str):
        text = value.strip()
        if text.startswith(("http://", "https://", "/")) and (
            re.search(r"\.(xlsx?|csv)(?:\?|$)", text, re.I)
            or "historical-archive" in text.lower()
        ):
            return text
    if isinstance(value, dict):
        # Prefer fields that commonly contain a download URL.
        for key in ("downloadUrl", "downloadURL", "url", "file", "href", "data"):
            if key in value:
                found = find_download(value[key])
                if found:
                    return found
        for item in value.values():
            found = find_download(item)
            if found:
                return found
    if isinstance(value, list):
        for item in value:
            found = find_download(item)
            if found:
                return found
    return None


def classify_response(response: requests.Response) -> tuple[bytes, str] | None:
    content = response.content
    content_type = (response.headers.get("content-type") or "").lower()
    disposition = (response.headers.get("content-disposition") or "").lower()
    if looks_like_xlsx(content) or "spreadsheet" in content_type or "excel" in content_type or ".xlsx" in disposition:
        return content, "xlsx"
    if "csv" in content_type or ".csv" in disposition:
        return content, "csv"
    return None


def download_archive() -> tuple[bytes, str]:
    s = session()
    errors: list[str] = []

    # The official page links directly to this archive endpoint. Add a cache
    # buster so an intermediate cache cannot keep yesterday's workbook.
    for attempt in range(1, 4):
        try:
            response = s.get(f"{API}&_={int(time.time())}", timeout=(20, 120), allow_redirects=True)
            response.raise_for_status()
            direct = classify_response(response)
            if direct:
                return direct
            try:
                download_url = find_download(response.json())
            except (ValueError, json.JSONDecodeError):
                download_url = None
            if download_url:
                downloaded = s.get(urljoin(API, download_url), timeout=(20, 120), allow_redirects=True)
                downloaded.raise_for_status()
                classified = classify_response(downloaded)
                if classified:
                    return classified
                raise RuntimeError(f"download URL returned unsupported content-type {downloaded.headers.get('content-type')}")
            errors.append(f"API attempt {attempt}: unsupported content-type {response.headers.get('content-type')}")
        except Exception as exc:  # noqa: BLE001 - aggregate official source failures
            errors.append(f"API attempt {attempt}: {exc}")
        time.sleep(attempt * 2)

    # Official-page fallback in case the API response contract changes.
    try:
        page = s.get(PAGE, timeout=(20, 90), allow_redirects=True)
        page.raise_for_status()
        links = re.findall(r'href=["\']([^"\']+(?:xlsx?|csv|historical-archive)[^"\']*)["\']', page.text, re.I)
        for link in links:
            downloaded = s.get(urljoin(PAGE, link), timeout=(20, 120), allow_redirects=True)
            if not downloaded.ok:
                continue
            classified = classify_response(downloaded)
            if classified:
                return classified
    except Exception as exc:  # noqa: BLE001
        errors.append(f"page fallback: {exc}")

    raise RuntimeError("Official SPDR archive download failed: " + " | ".join(errors[-6:]))


def extract_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    best: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        for header_index, row in enumerate(rows[:80]):
            headers = [norm(cell) for cell in row]
            date_index = next((i for i, text in enumerate(headers) if text in {"date", "as of date"} or text.endswith(" date")), None)
            tonnes_index = next((i for i, text in enumerate(headers) if "tonne" in text or "metric ton" in text), None)
            if date_index is None or tonnes_index is None:
                continue
            extracted: list[dict[str, Any]] = []
            for values in rows[header_index + 1 :]:
                if max(date_index, tonnes_index) >= len(values):
                    continue
                official_date = parse_date(values[date_index])
                try:
                    tonnes = float(str(values[tonnes_index]).replace(",", "").strip())
                except (TypeError, ValueError):
                    continue
                if official_date and 100 < tonnes < 5000:
                    extracted.append({"date": official_date, "holdings": round(tonnes, 3)})
            if len(extracted) > len(best):
                best = extracted
    return best


def extract_csv(content: bytes) -> list[dict[str, Any]]:
    rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace"))))
    best: list[dict[str, Any]] = []
    for header_index, row in enumerate(rows[:80]):
        headers = [norm(cell) for cell in row]
        date_index = next((i for i, text in enumerate(headers) if text in {"date", "as of date"} or text.endswith(" date")), None)
        tonnes_index = next((i for i, text in enumerate(headers) if "tonne" in text or "metric ton" in text), None)
        if date_index is None or tonnes_index is None:
            continue
        extracted: list[dict[str, Any]] = []
        for values in rows[header_index + 1 :]:
            if max(date_index, tonnes_index) >= len(values):
                continue
            official_date = parse_date(values[date_index])
            try:
                tonnes = float(values[tonnes_index].replace(",", "").strip())
            except (ValueError, AttributeError):
                continue
            if official_date and 100 < tonnes < 5000:
                extracted.append({"date": official_date, "holdings": round(tonnes, 3)})
        if len(extracted) > len(best):
            best = extracted
    return best


def existing_payload() -> dict[str, Any]:
    try:
        data = json.loads(OUT.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, ValueError):
        return {}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", help="Use a downloaded XLSX/CSV for local validation")
    args = parser.parse_args()

    if args.input_file:
        path = Path(args.input_file)
        content = path.read_bytes()
        kind = "xlsx" if looks_like_xlsx(content) or path.suffix.lower() in {".xlsx", ".xlsm"} else "csv"
    else:
        content, kind = download_archive()

    data = extract_csv(content) if kind == "csv" else extract_xlsx(content)
    if not data:
        raise RuntimeError("Could not locate official Date and Tonnes of Gold columns")

    unique = {row["date"]: row for row in data}
    records = sorted(unique.values(), key=lambda row: row["date"])[-30:]
    if len(records) < 2:
        raise RuntimeError("Official archive returned fewer than two valid records")

    latest = records[-1]
    previous = records[-2]
    daily_change = round(float(latest["holdings"]) - float(previous["holdings"]), 3)
    old = existing_payload()
    old_records = old.get("records") if isinstance(old.get("records"), list) else []
    checked_at = datetime.now(timezone.utc).isoformat()
    records_changed = old_records != records
    data_updated_at = checked_at if records_changed else (old.get("dataUpdatedAt") or old.get("updatedAt") or checked_at)

    # Persist the latest successful official connection even when SPDR has not
    # published a new record. This lets the website distinguish "checked and
    # unchanged" from "not connected" without altering the holdings history.
    payload = {
        "engineVersion": "etf-stable-2",
        "source": "SPDR Gold Shares official historical archive",
        "sourceUrl": PAGE,
        "archiveUrl": API,
        "updatedAt": data_updated_at,
        "dataUpdatedAt": data_updated_at,
        "checkedAt": checked_at,
        "lastSuccessfulConnectionAt": checked_at,
        "officialDate": latest["date"],
        "latestHoldings": latest["holdings"],
        "dailyChange": daily_change,
        "records": records,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    action = "Updated official holdings" if records_changed else "Official archive checked; no new record"
    print(
        f"{action}. Wrote {len(records)} records to {OUT}; latest {latest['date']} "
        f"{latest['holdings']:.3f} t ({daily_change:+.3f} t); checked {checked_at}"
    )


if __name__ == "__main__":
    main()
